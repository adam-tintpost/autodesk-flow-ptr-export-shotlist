
# ============================================================================
# CUSTOM TOOLKIT APP IMPLEMENTATION GUIDE
# ============================================================================
#
# This guide describes how to integrate ANY custom Toolkit app into
# Autodesk Flow Production Tracking.
#
# ---------------------------------------------------------------------------
# 1. CREATE THE CUSTOM APP
# ---------------------------------------------------------------------------
#
# Create a Toolkit app with the following structure:
#
# tk-your-app/
# ├── app.py
# ├── info.yml
# ├── python/
# │   └── __init__.py
# └── resources/
#     └── (icons, templates, etc.)
#
# app.py
#     Main application code.
#
# info.yml
#     Toolkit metadata (display name, supported engines, etc.)
#
# python/
#     Optional helper modules.
#
# resources/
#     Templates, icons and other bundled assets.
#
#
# ---------------------------------------------------------------------------
# 2. REGISTER THE APP LOCATION
# ---------------------------------------------------------------------------
#
# File:
#     config/env/includes/app_locations.yml
#
# Add a descriptor pointing to the application.
#
# Example (Git):
#
# apps.tk-your-app.location:
#   type: git
#   path: git@github.com:<username>/<repository>.git
#   version: v1.0.0
#
# Example (Path):
#
# apps.tk-your-app.location:
#   type: path
#   path: /path/to/tk-your-app
#
#
# ---------------------------------------------------------------------------
# 3. CREATE AN APP SETTINGS FILE
# ---------------------------------------------------------------------------
#
# Create:
#
# config/env/includes/settings/tk-your-app.yml
#
# Contents:
#
# includes:
# - ../app_locations.yml
#
# settings.tk-your-app:
#   location: "@apps.tk-your-app.location"
#
#
# ---------------------------------------------------------------------------
# 4. INCLUDE THE SETTINGS FILE
# ---------------------------------------------------------------------------
#
# File:
#
# config/env/includes/settings/tk-shotgun.yml
#
# Add the settings include:
#
# includes:
#   ...
#   - ./tk-your-app.yml
#
#
# ---------------------------------------------------------------------------
# 5. LOAD THE APP INTO THE DESIRED CONTEXT
# ---------------------------------------------------------------------------
#
# File:
#
# config/env/includes/settings/tk-shotgun.yml
#
# Add the app to whichever context(s) should display it:
#
# settings.tk-shotgun.project:
# settings.tk-shotgun.sequence:
# settings.tk-shotgun.shot:
# settings.tk-shotgun.asset:
# settings.tk-shotgun.task:
# etc.
#
# Example:
#
# settings.tk-shotgun.shot:
#   apps:
#     tk-your-app: "@settings.tk-your-app"
#
#
# ---------------------------------------------------------------------------
# 6. REGISTER THE TOOLKIT COMMAND
# ---------------------------------------------------------------------------
#
# Inside app.py:
#
# def init_app(self):
#
#     self.engine.register_command(
#         "your_command",
#         self.your_function,
#         {
#             "title": "Menu Name",
#             "supports_multiple_selection": True,
#         }
#     )
#
# This creates the menu item inside Flow Production Tracking.
#
#
# ---------------------------------------------------------------------------
# 7. RESTART TOOLKIT
# ---------------------------------------------------------------------------
#
# Restart the Flow Desktop application or reload Toolkit after making changes.
#
# ============================================================================


import os
import tempfile
import sgtk
import time

from datetime import datetime
from copy import copy
from concurrent.futures import ThreadPoolExecutor, as_completed
import urllib.request
import shutil

from PIL import Image as PILImage

from PySide2 import QtWidgets

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

class ShotlistExport(sgtk.platform.Application):


    # debug toggle
    DEBUG_MODE = False
    # to check if thumbnails are slowing down execution
    THUMBNAILS_DISABLED = False


    # CLASS ATTRIBUTES

    IMAGE_MAX_WIDTH = 240
    IMAGE_COLUMN_WIDTH = 30
    POINTS_PER_PIXEL_ASPECT = 0.75

    FRAMES_PER_SECOND = 30
    MAX_CONCURRENT_DOWNLOADS = 5

    SHOTCODE_COLUMN = 1
    THUMBNAIL_COLUMN = 2
    TIMECODE_COLUMN = 3
    VERSION_COLUMN = 4
    DESCRIPTION_COLUMN = 5
    HOURS_COLUMN = 6
    STATUS_COLUMN = 7

    TEMPLATE_ROW = 4

    # Is set at the beginning of the execution
    BASE_DIR = None
    THUMBNAIL_DIR = None
    TEMPLATE_FILE = None


    STATUS_INFO = {

        "act": {
            "name": "Active",
            "color": "19761B"
        },

        "apr": {
            "name": "Approved",
            "color": "B3B3B3"
        },

        "acr": {
            "name": "Awaiting Client Review",
            "color": "005C83"
        },

        "ca": {
            "name": "Client approved",
            "color": "A1EC9A"
        },

        "cap": {
            "name": "Client Approved",
            "color": "19761B"
        },

        "clsd": {
            "name": "Closed",
            "color": "969696"
        },

        "cmpt": {
            "name": "Complete",
            "color": "929292"
        },

        "cfrm": {
            "name": "Confirmed",
            "color": "A1EC9A"
        },

        "dlvr": {
            "name": "Delivered",
            "color": "D5D5D5"
        },

        "dis": {
            "name": "Disabled",
            "color": "CC0001"
        },

        "fdbk": {
            "name": "Feedback to adress",
            "color": "B700BC"
        },

        "fta": {
            "name": "Feedback to adress",
            "color": "DEB6FF"
        },

        "fin": {
            "name": "Final",
            "color": "969696"
        },

        "ing": {
            "name": "In grade",
            "color": "006EFC"
        },

        "ion": {
            "name": "In_Online",
            "color": "0295D8"
        },

        "ip": {
            "name": "In Progress",
            "color": "CAE1CA"
        },

        "inp": {
            "name": "In progress",
            "color": "FFFFFF"
        },

        "ia": {
            "name": "Internally approved",
            "color": "98F2FB"
        },

        "iap": {
            "name": "Internally approved",
            "color": "CAEDC5"
        },

        "na": {
            "name": "N/A",
            "color": "FFFFFF"
        },

        "omt": {
            "name": "Omit",
            "color": "E17F81"
        },

        "omit": {
            "name": "Omit",
            "color": "FE9798"
        },

        "hld": {
            "name": "On Hold",
            "color": "969696"
        },

        "onhold": {
            "name": "On hold",
            "color": "FECD8A"
        },

        "opn": {
            "name": "Open",
            "color": "D5D5D5"
        },

        "pndng": {
            "name": "Pending",
            "color": "969696"
        },

        "rev": {
            "name": "Pending Review",
            "color": "95E3A7"
        },

        "pr": {
            "name": "Pending Review",
            "color": "EFFBCB"
        },

        "prw": {
            "name": "Pending review",
            "color": "004A00"
        },

        "pnrw": {
            "name": "Pending Review",
            "color": "BCDAFC"
        },

        "qc": {
            "name": "QC Ready For Delivery",
            "color": "016E01"
        },

        "rts": {
            "name": "Ready to start",
            "color": "FBFDCC"
        },

        "recd": {
            "name": "Received",
            "color": "FFFFFF"
        },

        "res": {
            "name": "Resolved",
            "color": "969696"
        },

        "stg": {
            "name": "Send To Grade",
            "color": "FFFFFF"
        },

        "supap": {
            "name": "supervisor approved",
            "color": "FFFD1C"
        },

        "too": {
            "name": "To_Online",
            "color": "FFFFFF"
        },

        "toooo": {
            "name": "To_Online",
            "color": "FDAFFB"
        },

        "vwd": {
            "name": "Viewed",
            "color": "929292"
        },

        "wtg": {
            "name": "Waiting to Start",
            "color": "D5D5D5"
        },
        "rfgi": {
            "name": "Ready For Grade Insert",
            "color": "FFFFFF"
        }
    }

    def init_app(self):

        if self.DEBUG_MODE:
            self.log_info(">>> tk-shotlist-export INIT SUCCESS <<<")

        props = {
            "title": "Generate Excel Shotlist",
            "supports_multiple_selection": True,
        }

        self.engine.register_command(
            "shotlist_export",
            self.export_shotlist,
            props,
        )

    def millisec_to_tc(self, milliseconds):

        if milliseconds == "" or milliseconds is None:
            return ""

        total_seconds = milliseconds / 1000.0

        hours = int(total_seconds // 3600)
        total_seconds %= 3600

        minutes = int(total_seconds // 60)
        total_seconds %= 60

        seconds = int(total_seconds)

        fractional = total_seconds - seconds
        frames = round(fractional * self.FRAMES_PER_SECOND)

        if frames >= self.FRAMES_PER_SECOND:
            frames = 0
            seconds += 1

        return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

    def copy_row_format(self, ws, source_row, target_row):

        for col in range(1, ws.max_column + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)

            target._style = copy(source._style)
    
    def download_thumbnail(self, shot, start_time):
        shot_name = shot["code"]
        thumbnail_url = shot.get("image")

        if not thumbnail_url:
            return shot_name, None

        if self.DEBUG_MODE:
            self.log_info(f"Starting download of thumbnail for {shot_name}")

        local_img = os.path.join(
            self.THUMBNAIL_DIR,
            f"{shot_name}.jpg"
        )

        try:
            urllib.request.urlretrieve(
                thumbnail_url,
                local_img
            )
        except Exception as e:
            self.log_error(f"Thumbnail download failed for {shot_name}: {e}")
            
            return shot_name, None

        if self.DEBUG_MODE:
            self.log_time(f"Finished downloading thumbnail for {shot_name}", start_time)

        # resize image immediately to save processing time in excel
        with PILImage.open(local_img) as img:
            img.thumbnail((self.IMAGE_MAX_WIDTH, 9999), PILImage.Resampling.LANCZOS)
            img.save(local_img, quality=90)

        return shot_name, local_img

    def log_time(self, label, start_time):
        if self.DEBUG_MODE:
            elapsed = time.time() - start_time
            self.log_info(f"{label}: {elapsed:.3f}s")

    def export_shotlist(self, entity_type, entity_ids):

        start_time = time.time()

        if self.DEBUG_MODE:
            self.log_info("::::: DEBUG MODE ACTIVE :::::")

        sg = self.engine.shotgun

        self.BASE_DIR = os.path.dirname(__file__)

        self.THUMBNAIL_DIR = tempfile.mkdtemp(prefix="shotlist_export_")

        self.TEMPLATE_FILE = os.path.join(
            self.BASE_DIR,
            "resources",
            "SHOTLIST_TEMPLATE.xlsx",
        )

        if self.DEBUG_MODE:
            self.log_info(f"Entity type: {entity_type}")
            self.log_info(f"Entity IDs: {entity_ids}")

        # ----------------------------
        # MAIN EXPORT
        # ----------------------------
        if self.DEBUG_MODE:
            self.log_time("Starting shotlist export", start_time)


        shot_ids = entity_ids

        # retrieve shots
        shots = sg.find(
            "Shot",
            [["id", "in", shot_ids]],
            [
                "id",
                "code",
                "sg_status_list",
                "image",
                "sg_tc_in",
                "description",
                "sg_sequence",
                "sg_latest_version",
                "smart_duration_summary_display"
            ],
            order=[{"field_name": "code", "direction": "asc"}],
        )

        if self.DEBUG_MODE:
            self.log_time("Retrieved %d shots" % len(shots), start_time)


        # ----------------------------
        # CREATE EXCEL
        # ----------------------------

        try:

            wb = load_workbook(self.TEMPLATE_FILE)
            ws = wb.active
            ws.column_dimensions["B"].width = self.IMAGE_COLUMN_WIDTH

            row = 4 

            # CREATE HEADER INFO
            
            project = sg.find_one(
                "Project",
                [["id", "is", self.context.project["id"]]],
                ["name"],
            )

            project_name = project["name"]

            current_user = sgtk.get_authenticated_user()

            user = self.shotgun.find_one(
                "HumanUser",
                [["login", "is", current_user.login]],
                ["name"]
            )

            # Sets the name of the excelsheet
            ws.title = project_name

            # Fill header project info
            ws["A1"] = project_name
            ws["B2"] = datetime.now().strftime("%Y-%m-%d")
            user_name = user["name"]
            ws["C2"] = user_name

            # Retrieve task-info
            tasks = sg.find(
                "Task",
                [
                    ["entity", "type_is", "Shot"],
                    ["entity", "in", [{"type": "Shot", "id": sid} for sid in shot_ids]]

                ],
                [
                    "entity",
                    "content",
                    "duration"
                ]
            )

            if self.DEBUG_MODE:
                self.log_time("Retrieved tasks for shots", start_time)

            task_durations = {} 

            for task in tasks:
                shot_id = task["entity"]["id"]

                duration = task.get("duration") or 0

                if shot_id not in task_durations:
                    task_durations[shot_id] = 0

                task_durations[shot_id] += duration       

            if self.DEBUG_MODE:
                self.log_time("Calculated and summarized duration for shot-tasks", start_time)

            # Download thumbnails concurrently for efficiency

            thumbnail_results = {}

            with ThreadPoolExecutor(max_workers=self.MAX_CONCURRENT_DOWNLOADS) as executor:
                futures = [
                    executor.submit(self.download_thumbnail, shot, start_time)
                    for shot in shots
                ]

                for future in as_completed(futures):
                    shot_name, local_img = future.result()
                    thumbnail_results[shot_name] = local_img

            if self.DEBUG_MODE:
                self.log_time("Downloaded all thumbnails", start_time)

            for s in shots:
                code = s.get("code")
                status_code = s.get("sg_status_list")
                latest_version = s.get("sg_latest_version")
                tc_in = s.get("sg_tc_in")
                description = s.get("description")
                shot_id = s["id"]
                duration_minutes = task_durations[shot_id]

                # Formatting data
                tc_in_formatted = self.millisec_to_tc(tc_in)
                duration_hours = round(duration_minutes / 60, 1)

                # Get thumbnails from concurrent downloaded files
                local_img = thumbnail_results.get(s.get("code"))
                
                # Get status name and color hexadecimal

                status_info = self.STATUS_INFO.get(status_code)

                status_name = ""
                status_color = ""

                if status_info is None:
                    status_name = "Unavailable status"
                    status_color = "FFFFFF"
                else:
                    status_name = status_info["name"]
                    status_color = status_info["color"]

            
                ws.cell(row=row, column=self.SHOTCODE_COLUMN,       value=code)
                ws.cell(row=row, column=self.TIMECODE_COLUMN,       value=tc_in_formatted)
                ws.cell(row=row, column=self.VERSION_COLUMN,        value=latest_version)
                ws.cell(row=row, column=self.DESCRIPTION_COLUMN,    value=description)
                ws.cell(row=row, column=self.HOURS_COLUMN,          value=duration_hours)   
             
                # Set row style to match first
                self.copy_row_format(ws, self.TEMPLATE_ROW, row)

                
                # status cell both name and color set
                status_cell = ws.cell(row=row, column=self.STATUS_COLUMN, value=status_name)
                status_cell.fill = (PatternFill(fill_type="solid",fgColor=status_color))


                if local_img and os.path.exists(local_img) and not self.THUMBNAILS_DISABLED:

                    img = Image(local_img)

                    ws.add_image(img, f"B{row}")

                    ws.row_dimensions[row].height = img.height * self.POINTS_PER_PIXEL_ASPECT
                row += 1

            if self.DEBUG_MODE:
                self.log_time("Added all shots to excel file", start_time)


            # freeze top 3 header rows
            ws.freeze_panes = "A3"

            # ----------------------------
            # SAVE TEMP FILE
            # ----------------------------

            project_name = self.context.project["name"]

            # clean name (remove spaces and weird characters)
            project_name = project_name.replace(" ", "_")

            date = datetime.now().strftime("%Y%m%d")

            filename = f"Shotlist_{project_name}_{date}.xlsx"
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            default_filepath = os.path.join(desktop, filename)

            if self.DEBUG_MODE:
                self.log_time("Before saving workbook", start_time)

            # ----------------------------
            # RETURN FILE TO USER (DOWNLOAD)
            # ----------------------------

            save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
                None,
                "Save Shotlist",
                default_filepath,
                "Excel (*.xlsx)"
            )

            if self.DEBUG_MODE:
                self.log_time("After choosing save path and filename", start_time)

            if save_path:
                wb.save(save_path)

            if self.DEBUG_MODE:
                self.log_time("Workbook saved to file", start_time)

        # clean up the thumbnail directory before exiting program
        finally:
            shutil.rmtree(self.THUMBNAIL_DIR, ignore_errors=True)
            if self.DEBUG_MODE:
                self.log_time(f"Deleted temporary saved thumbnails: {self.THUMBNAIL_DIR}", start_time)
                self.log_info(f"Directory exists after cleanup: {os.path.exists(self.THUMBNAIL_DIR)}"
)


        if self.DEBUG_MODE:
            self.log_info("Saved to %s" % save_path)
            self.log_time("Total time for execution", start_time)

